from flask import render_template, redirect, request, url_for
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from .models import Clients, Communes, Institutions, Campaigns, InstitutionsCampaigns, SendMails, SendMailsCC, SendMailsFrom, Users, LinkTypes, ConnectionTypes, Links, ClientsDirectors, ClientsCategory, Regions, Provinces, LocalContact, GroupsMails, ItemGroupsMails, ClientsGroupsMails, GroupMailsSendMails

from app.logs.models import LogCampaignStatus, LogCommunesName, LogLocalContact, LogInstitutionsStatus

from . import admin_bp

from .forms import EditGroupsMailsForm, RegisterClientsForm, RegisterInstitutionsForm, ImportInstitutionsForm, RegisterCampaignsForm, RegisterUsersForm, UpdateUserForm, UpdatePasswordForm, RegisterLinkTypesForm, RegisterConnectionTypesForm, RegisterLinksForm, RegisterClientsCategoriesForm, ImportLinksForm, EditCommunesForm, EditInstitutionsForm, EditLinksForm, ImportLocalContactsForm, EditLocalContactForm, RegisterGroupsMailsForm, RegisterItemGroupsMailsForm, RegisterClientsGroupsMailsForm

from openpyxl import load_workbook
import pandas as pd
from os.path import join, dirname, realpath
from datetime import datetime
from app import login_manager, app, db
from app.decorators import admin_required
import io, pytz, pprint

@login_manager.user_loader
def load_user(user_id):
	return Users.query.get(int(user_id))

#global variables
data_xls_import = []
form_client = None
tmp_data_import = []
links_tmp_data_import = []

UPLOADS_PATH = join(dirname(realpath(__file__)), 'uploads/')

# Active item menu
institutions_module = None
clients_module = None
communes_module = None
campaigns_module = None
users_module = None
link_type_module = None
connection_type_module = None
link_module= None

def datefilter(value):
	tz = pytz.timezone('America/Santiago') # timezone you want to convert to from UTC
	utc = pytz.timezone('UTC')  
	value = utc.localize(value)
	local_dt = value.astimezone(tz)
	final_value= local_dt.strftime("%Y-%m-%d")

	return final_value
app.jinja_env.filters['datefilter'] = datefilter

def timefilter(value):
	tz = pytz.timezone('America/Santiago') # timezone you want to convert to from UTC
	utc = pytz.timezone('UTC')  
	value = utc.localize(value)
	local_dt = value.astimezone(tz)
	final_value= local_dt.strftime("%H:%M")

	return final_value
app.jinja_env.filters['timefilter'] = timefilter

# Administración - Clientes

@admin_bp.route("/admin/clientes")
@admin_bp.route("/admin/clientes/<string:client>", methods=["POST"])
@login_required
def clients(client = None):
	title = "Sistema de respaldos - Clientes"
	clients_module = True
	clients = Clients.query.join(ClientsDirectors, isouter = True).join(Users, isouter = True).join(ClientsCategory, isouter = True) \
		.add_column(Clients.id) \
		.add_column(Clients.name) \
		.add_column(ClientsCategory.name) \
		.add_column(Users.name) \
		.add_column(Users.lastname) \
		.order_by(Clients.name)

	client_delete_name = None
	if client is not None:
		client_delete_name = client
	return render_template(
		"clients.html", 
		title = title, 
		clients_module = clients_module,
		clients = clients,
		client_delete_name = client_delete_name)

@admin_bp.route("/admin/clients/registrar", methods=['POST','GET'])
@login_required
def clients_register():
	title = "Sistema de respaldos - Clientes"
	clients_module = True
	form = RegisterClientsForm()
	if form.validate_on_submit():
		name = form.name.data
		category = form.categories.data
		director = form.directors.data
		clients = Clients(name=name, category = category)
		clients.save()
		if director is not None:
			cd = ClientsDirectors()
			cd.clients_id = clients.id
			cd.directors= director
			cd.save()
		return redirect(url_for('admin.clients'))
	return render_template(
		"clients_register.html", 
		title = title, 
		clients_module = clients_module,
		form = form)

@admin_bp.route("/admin/clientes/eliminar/<int:id>", methods=["POST"])
@login_required
def clients_delete(id):
	if request.method == "POST":
		clients = Clients()
		client = clients.get_by_id(id = id)
		client.delete()
	return redirect(url_for('admin.clients', client = client.name), code=307)

@admin_bp.route("/admin/clientes/editar/<int:id>", methods=["GET", "POST"])
@login_required
def client_update_form(id):

	# lista de directores de servicios
	users = Users()
	users = users.get_directors()

	# datos del cliente
	client = Clients()
	client = client.get_by_id(id)

	# categorias
	categories = ClientsCategory()
	categories = categories.get_categories()

	director = client.get_director_by_client(id)

	if request.method == "POST":

		client.category_id = request.form.get('category')
		client.save()

		cd = ClientsDirectors()
		cd.clients_id = id
		cd.directors_id = request.form.get('user')
		cd.save()

		return redirect(url_for("admin.client_update_form", id = id))

	return render_template(
		"clients_detail.html",
		title = "Editar cliente - Sistema de respaldos",
		clients_module = True,
		users = users,
		director = director,
		client = client,
		categories = categories
	)

# Administración - Clientes - Categorias

@admin_bp.route("/admin/clientes/categorias", methods=["GET"])
@login_required
def clients_categories():
	categories = ClientsCategory.query.all()

	return render_template(
		"clients_categories.html",
		title = "Categorias - Clientes - Sistema de respaldos",
		clients_module = True,
		categories = categories)

@admin_bp.route("/admin/clientes/categorias/registrar", methods=["GET", "POST"])
@login_required
def clients_categories_register():
	form = RegisterClientsCategoriesForm()

	if form.validate_on_submit():
		name = form.name.data
		category = ClientsCategory(name=name)
		category.save()
		return redirect(url_for('admin.clients_categories'))

	return render_template(
		"clients_category_register.html",
		title = "Nueva categoria - Clientes - Sistema de respaldos",
		clients_module = True,
		form = form)

# Administración - Establecimientos

@admin_bp.route("/admin/establecimientos")
@admin_bp.route("/admin/establecimientos/<string:institution>", methods=["POST"])
@login_required
def institutions(institution = None):
	title = "Sistema de respaldos - Establecimientos"
	institutions_module = True
	institutions = Institutions.query.filter_by(active=True).order_by(Institutions.name)
	institution_delete_name = None
	if institution is not None:
		institution_delete_name = institution
	return render_template(
		"institutions.html", 
		title = title, 
		institutions_module = institutions_module,
		institutions = institutions,
		institution_delete_name= institution_delete_name)

@admin_bp.route("/admin/establecimientos/registrar", methods=['POST','GET'])
@login_required
def institutions_register():
	title = "Sistema de respaldos - Establecimientos"
	institutions_module = True
	form = RegisterInstitutionsForm()
	if form.validate_on_submit():
		name = form.name.data
		address = form.address.data
		commune = form.commune.data
		client = form.client.data
		institution = Institutions(name=name, address=address, commune=commune, client=client)
		institution.save()
		return redirect(url_for('admin.institutions'))
	return render_template(
		"institutions_register.html", 
		title = title, 
		institutions_module = institutions_module,
		form = form)

@admin_bp.route("/admin/establecimientos/editar/<int:id>", methods=['GET','POST'])
@login_required
def institutions_edit(id):
	form = EditInstitutionsForm()

	if form.validate_on_submit():

		institution = Institutions.query.filter_by(id=form.institution_id.data).first()
		institution.name = form.name.data
		institution.commune = form.commune.data
		institution.address = form.address.data
		msg = institution.save()

		return render_template(
		"institutions_edit.html", 
		title = "Editar - Establecimientos - Sistema de respaldos", 
		institutions_module = True,
		success_msg = "Se han guardado los cambios exitosamente.",
		form = form,
		id = id)

	institution = Institutions.query.filter_by(id=id).first()

	form.institution_id.data = institution.id
	form.name.data = institution.name
	form.address.data = institution.address
	form.commune.data = institution.commune
	form.client.data = institution.client

	return render_template(
		"institutions_edit.html", 
		title = "Editar - Establecimientos - Sistema de respaldos", 
		institutions_module = True,
		form = form,
		id=id)

@admin_bp.route("/admin/establecimientos/eliminar/<int:id>", methods=["POST"])
@login_required
def institution_delete(id):
	if request.method == "POST":
		institutions = Institutions()
		institution = institutions.get_by_id(id = id)
		institution.update_active()
	return redirect(url_for('admin.institutions', institution = institution.name), code=307)

@admin_bp.route("/admin/establecimientos/importar", methods=["GET", "POST"])
@login_required
def institutions_import():

	form = ImportInstitutionsForm()

	return render_template(
		"institutions_import.html", 
		title = "Importar establecimientos - Sistema de respaldos",
		form=form,
		institutions_module = True)

@admin_bp.route("/admin/establecimientos/importar/preview", methods=["POST"])
@login_required
def institutions_import_preview():

	form = ImportInstitutionsForm()
	global form_client
	count_institutions = 0

	if form.validate_on_submit():
		filename = secure_filename(form.upload.data.filename)
		global data_xls_import 
		global tmp_data_import

		form_client = form.client.data

		form.upload.data.save(UPLOADS_PATH+filename)

		tmp_data_import = load_workbook(UPLOADS_PATH+filename, data_only=True)
		data_xls_import = pd.read_excel(UPLOADS_PATH+filename)

		count_institutions = len(data_xls_import)

		data_xls_import_html = data_xls_import.to_html(classes="table table-bordered")

	return render_template(
		"institutions_import.html", 
		form=form, 
		data_xls_import=data_xls_import_html, 
		client = form_client, 
		count_institutions = count_institutions,
		institutions_module = True)

@admin_bp.route("/admin/establecimientos/importar/confirmar", methods=["POST"])
@login_required
def institutions_import_confirm():
	
	global tmp_data_import
	global form_client

	form = ImportInstitutionsForm()

	repeat = True
	count = 1
	num_rows = 0
	count_institutions_repeat = 0
	institutions_repeated = []
	institutions_no_commune = []
	institutions_registered = 0
	exist = False
	import_finished = False
	client = None
	
	while repeat == True:
		num_rows = num_rows + 1
		count = count + 1

		letra_a = "A"+str(count)
		letra_b = "B"+str(count)
		letra_c = "C"+str(count)

		celda_a = tmp_data_import['establecimientos'][letra_a].value
		celda_b = tmp_data_import['establecimientos'][letra_b].value
		celda_c = tmp_data_import['establecimientos'][letra_c].value

		# se valida si quedan más registros
		if (celda_a == None and celda_b == None and celda_c == None):
			repeat = False
		else:
			# validación de comuna
			c = Communes()
			commune_exist = c.validate_exist(name = str(celda_c))
			communes = c.get_by_name(name = str(celda_c))
			cl = Clients()
			client = cl.get_id_by_name(name = str(form_client))

			if commune_exist is not False:
				# validacion para evitar que el establecimiento se repita
				institution = Institutions()
				institution = institution.validate_exist_per_client(name=str(celda_a), client=client.id)

				if institution is False:
					institution = Institutions(name=str(celda_a), address=str(celda_b), commune_id=commune_exist, client_id=client.id)
					institution.save()
					institutions_registered = institutions_registered + 1
				else:
					count_institutions_repeat = count_institutions_repeat + 1
					institutions_repeated.append(celda_a)
			# la comuna no existe, se registra
			else:
				elements = {'name': str(celda_a), 'commune':celda_c}
				institutions_no_commune.append(elements)
	import_finished = True

	return render_template(
		"institutions_import.html",
		client = client,
		form=form, 
		institutions_repeated=institutions_repeated, 
		institutions_no_commune=institutions_no_commune,
		import_finished = import_finished, 
		institutions_module = True,
		institutions_registered = institutions_registered)

@admin_bp.route("/admin/establecimientos/contactos_locales", methods=["GET"])
@login_required
def local_contacts(client = None):

	local_contacts = LocalContact.query.filter(LocalContact.active == True)

	return render_template(
		"local_contacts.html",
		title = "Contactos locales - Sistema de respaldos",
		institutions_module = True,
		local_contacts = local_contacts
	)

@admin_bp.route("/admin/establecimientos/contactos_locales/eliminar", methods=["POST"])
@login_required
def local_contact_delete():

	id = request.form.get("id")
	local_contacts = LocalContact.query.filter(LocalContact.active == True, LocalContact.id == id).first()
	local_contacts.delete()

	return redirect(url_for("admin.local_contacts"))

@admin_bp.route("/admin/establecimientos/contactos_locales/editar/<int:id>", methods=["GET","POST"])
@login_required
def local_contacts_edit(id):
	msg_success_update = False

	local_contacts = LocalContact.query.filter_by(id = id).first()
	logs = LogLocalContact.query.filter(LogLocalContact.local_contact_id == id).order_by(LogLocalContact.id.desc()).all()

	form = EditLocalContactForm()

	if form.validate_on_submit():
		local_contacts.name = form.name.data
		local_contacts.phone = form.phone.data
		local_contacts.email = form.email.data
		local_contacts.save()

		msg_success_update = True

		local_contacts = LocalContact.query.filter_by(id = id).first()

	form.institution.data = str(local_contacts.institution)
	form.name.data = local_contacts.name
	form.phone.data = local_contacts.phone
	form.email.data = local_contacts.email

	active = local_contacts.active

	return render_template(
		"local_contacts_edit.html",
		title = "Contactos locales - Sistema de respaldos",
		institutions_module = True,
		form = form,
		active = active,
		local_contact_id = id,
		logs = logs,
		msg_success_update = msg_success_update
	)

@admin_bp.route("/admin/establecimientos/contactos_locales/activar", methods=["POST"])
@login_required
def local_contact_activate():

	id = request.form.get("local_contact_id")
	l = LocalContact.query.filter(LocalContact.id == id).first()
	l.activate()

	return redirect(url_for("admin.local_contacts_edit", id=id))

@admin_bp.route("/admin/establecimientos/contactos_locales/desactivar", methods=["POST"])
@login_required
def local_contact_deactivate():

	id = request.form.get("local_contact_id")
	l = LocalContact.query.filter(LocalContact.id == id).first()
	l.deactivate()

	return redirect(url_for("admin.local_contacts_edit", id=id))

@admin_bp.route("/admin/establecimientos/contactos_locales/importar", methods=["GET","POST"])
@login_required
def local_contacts_import():

	form = ImportLocalContactsForm()

	return render_template(
		"local_contacts_import.html",
		title = "Contactos locales - Sistema de respaldos",
		institutions_module = True,
		form = form
	)

@admin_bp.route("/admin/establecimientos/contactos_locales/preview", methods=["POST"])
@login_required
def local_contacts_preview():

	form = ImportLocalContactsForm()

	global form_client
	global form_client_id
	global data_xls_import 
	global tmp_data_import
	count_contacts = 0
	data_xls_import_html = None

	if form.validate_on_submit():

		filename = secure_filename(form.upload.data.filename)
		form_client = form.client.data

		form.upload.data.save(UPLOADS_PATH+filename)

		tmp_data_import = load_workbook(UPLOADS_PATH+filename, data_only=True)
		data_xls_import = pd.read_excel(UPLOADS_PATH+filename, keep_default_na=False)

		count_contacts = len(data_xls_import)

		data_xls_import_html = data_xls_import.to_html(classes="table table-bordered")
	
	return render_template(
		"local_contacts_import.html", 
		form=form, 
		data_xls_import=data_xls_import_html, 
		client = form_client, 
		count_contacts = count_contacts,
		institutions_module = True)

# esto se ejecuta al confirmar la importación
# se registran los datos de la planilla a la base de datos
@admin_bp.route("/admin/contactos_locales/importar/confirmar", methods=["POST"])
@login_required
def local_contacts_import_confirm():
	
	global tmp_data_import
	global form_client

	form = ImportLocalContactsForm()

	client = Clients()
	client = client.get_id_by_name(name = str(form_client))

	repeat = True
	count = 1
	num_rows = 0

	local_contact_registered = 0
	import_finished = False
	
	no_institution = []
	element_no_institution = []

	while repeat == True:
		num_rows = num_rows + 1
		count = count + 1

		letra_a = "A"+str(count)
		letra_b = "B"+str(count)
		letra_c = "C"+str(count)
		letra_d = "D"+str(count)

		celda_a = tmp_data_import['contactos_locales'][letra_a].value
		celda_b = tmp_data_import['contactos_locales'][letra_b].value
		celda_c = tmp_data_import['contactos_locales'][letra_c].value
		celda_d = tmp_data_import['contactos_locales'][letra_d].value

		# se valida si quedan más registros
		if (celda_a == None and celda_b == None and celda_c == None and celda_d == None):
			repeat = False
		else:
			

			# validación de establecimiento
			institution_exist = Institutions()
			institution_name = institution_exist.validate_by_name(str(celda_a))
			institution_exist = institution_exist.get_institution_by_name(name=str(institution_name), client=client.id)
					
			if institution_exist is None:
				# enlaces con establecimientos no encontrados
				element_no_institution = {
					'institution': str(celda_a), 
					'name': str(celda_b),
					'phone': str(celda_c),
					'email': str(celda_d)
					}
				no_institution.append(element_no_institution)

			else:
				local_contact = LocalContact(name = str(celda_b), phone = str(celda_c), email = celda_d, institution = institution_exist)
				local_contact.save()
				local_contact_registered = local_contact_registered + 1

	import_finished = True

	return render_template(
		"local_contacts_import.html",
		form=form, 
		import_finished = import_finished, 
		links_module = True,
		client = client,
		local_contact_registered = local_contact_registered,
		no_institution = no_institution)

# Administración - Regiones
@admin_bp.route("/admin/regiones")
@login_required
def regions(regions = None):
	regions = Regions.query.order_by(Regions.name).all()
	return render_template(
		"regions.html", 
		title = "Regiones - Sistema de respaldos", 
		regions_module = True,
		regions = regions)

# Administración - Provincias
@admin_bp.route("/admin/provincias")
@login_required
def provinces(provinces = None):
	provinces = Provinces.query.order_by(Provinces.name).all()
	return render_template(
		"provinces.html", 
		title = "Provincias - Sistema de respaldos", 
		provinces_module = True,
		provinces = provinces)

# Administración - Comunas
@admin_bp.route("/admin/comunas")
@admin_bp.route("/admin/comunas/<string:commune>", methods=["POST"])
@login_required
def communes(commune = None):
	title = "Sistema de respaldos - Comunas"
	communes_module = True
	communes = Communes.query.order_by(Communes.name).all()
	commune_delete_name = None
	if commune is not None:
		commune_delete_name = commune
	return render_template(
		"communes.html", 
		title = title, 
		communes_module = communes_module,
		communes = communes, 
		commune_delete_name = commune_delete_name)

@admin_bp.route("/admin/comunas/editar/<int:id>", methods=['GET','POST'])
@login_required
def communes_edit(id):

	# comunas
	commune = Communes.query.filter_by(id=id).first()

	logs = LogCommunesName.query.filter_by(commune_id = id).order_by(LogCommunesName.id.desc()).all()

	form = EditCommunesForm()

	if form.validate_on_submit():
		if form.name_after.data != form.name_before.data:
			commune = Communes()
			commune.id = id
			commune.name = form.name_after.data
			commune.setName(name_before = form.name_before.data)
			logs = LogCommunesName.query.filter_by(commune_id = id).order_by(LogCommunesName.id.desc()).all()

	form.name_before.data = commune.name
	form.name_after.data = commune.name

	return render_template(
		"communes_edit.html",
		title = "Editar comuna - Sistema de respaldos",
		communes_module = True,
		commune = commune,
		form = form,
		commune_id = id,
		logs = logs
	)

@admin_bp.route("/admin/comunas/eliminar/<int:id>", methods=["POST"])
@login_required
def communes_delete(id):
	if request.method == "POST":
		communes = Communes()
		commune = communes.get_by_id(id = id)
		commune.delete()
	return redirect(url_for('admin.communes', commune = commune.name), code=307)

# Administración camapañas

# lista de campañas
@admin_bp.route("/admin/campanias")
@login_required
def campaigns():

	title = "Sistema de respaldos - Campañas"
	campaigns_module = True
	campaigns = Campaigns.query.join(LogCampaignStatus, isouter = True) \
		.join(Clients, isouter = True) \
		.join(Users, isouter = True) \
		.add_column(Campaigns.id) \
		.add_column(Campaigns.name) \
		.add_column(Campaigns.client) \
		.add_column(Campaigns.period) \
		.add_column(Campaigns.active) \
		.add_column(Campaigns.created.label("campaign_created")) \
		.add_column(Clients.name.label("client_name")) \
		.add_column(LogCampaignStatus.user.label("finish_user")) \
		.add_column(Users.name.label("user_name")) \
		.add_column(Users.lastname.label("user_lastname")) \
		.add_column(LogCampaignStatus.created).order_by(Campaigns.id, Campaigns.name)
	return render_template(
		'campaigns.html',
		campaigns_module=campaigns_module,
		title=title,
		campaigns = campaigns)

# establecimientos por campaña
@admin_bp.route("/admin/campanias/<int:id>/establecimientos", methods=['GET', 'POST'])
@login_required
def campaigns_institutions(id):

	msg_success = None
	title = "Sistema de respaldos - Campañas"
	campaigns_module = True

	# establecimientos por campaña
	institutions = Institutions()
	institutions = institutions.get_institutions_by_campaigns(id)

	# identificar cliente según la campaña
	client = Clients()
	client = client.get_client_by_campaign(id = id)

	if request.method == 'POST':
		
		institution_name = request.form.get('institution_name')

		# se ejecuta si se insertan establecimientos
		if request.form.get('insert_institutions'):
			institutions = request.form.getlist('chk_institutions')
			campaign = request.form.get('campaign')

			for institution in institutions:
				institution = InstitutionsCampaigns(institution_id = institution, campaign_id = campaign)
				institution.save()
				# el establecimiento al ser insertada a una campaña automáticamente se actualiza a estado = 1 (en gestión)
				institution_update = Institutions()
				institution_update.id = institution.institution_id
				institution_update.status_id = 1
				institution_update.set_status(campaign_institution_id=institution.id)
			
			updated_institutions = Institutions()
			institutions = updated_institutions.get_institutions_by_campaigns(campaign)

			msg_success = "Se han insertado el/los establecimientos correctamente."
			return render_template(
			'campaigns_institutions.html',
			campaigns_module=campaigns_module,
			title=title,
			institutions = institutions,
			client = client,
			campaign_id = id,
			msg_discard_success = msg_success)
		
		else:
			campaign_id = request.form.get('campaign_id')
			institution_id = request.form.get('institution_id')
			try:
				row = InstitutionsCampaigns.query.filter(InstitutionsCampaigns.institution_id == institution_id, campaign_id == campaign_id).first()
				i = Institutions()
				i.finish_insts(institution_id = institution_id, campaign_institution_id = row.id, current_user_id=current_user.id, discard=True)
			except Exception as e:
				print(e)
			finally:
				updated_institutions = Institutions()
				institutions = updated_institutions.get_institutions_by_campaigns(id)
				msg_success = "Se ha quitado el establecimiento <b>"+institution_name+"</b> de la campaña exitosamente."

	return render_template(
		'campaigns_institutions.html',
		campaigns_module=campaigns_module,
		title=title,
		institutions = institutions,
		client = client,
		campaign_id = id,
		msg_discard_success = msg_success)

@admin_bp.route("/admin/campanias/registrar", methods=['POST','GET'])
@admin_bp.route("/admin/campanias/registrar/<string:error>", methods=['POST','GET'])
@login_required
@admin_required
def campaigns_register(error = None):

	title = "Sistema de respaldos - Campañas"
	campaigns_module = True
	form = RegisterCampaignsForm()

	form_year_min = int(datetime.now().year) - 1
	form_year_max = int(datetime.now().year) + 1
	
	return render_template(
		'campaigns_register.html',
		title = title,
		campaigns_module = campaigns_module,
		form = form,
		form_year_min = form_year_min,
		form_year_max = form_year_max,
		error_msg = error
	)

@admin_bp.route("/admin/campanias/seleccion_establecimientos", methods=['POST'])
@admin_bp.route("/admin/campanias/seleccion_establecimientos/<int:selected_client_id>", methods=['GET','POST'])
@login_required
def campaigns_select_institutions(selected_client_id = None):

	title = "Sistema de respaldos - Campañas"
	campaigns_module = True
	form = RegisterCampaignsForm()
	client = form.client.data
	period = form.period.data
	error_msg = None
	campaign_id = None

	if selected_client_id == None:

		# se comprueba si existe una campaña con el mismo nombre
		campaigns = Campaigns()
		campaign_exists = campaigns.exist(campaign = str(form.name.data))

		if campaign_exists is False:
			campaign = Campaigns(name=str(form.name.data), client = client, period = period)
			campaign.save()
			campaign_id = campaign.last_id()
		else:
			# ya existe camapaña con el mismo nombre asi que se redirecciona al formulario de registro
			error_msg = True
			return redirect(url_for('admin.campaigns_register', error = error_msg))

		# se obtienen establecimientos que no estén en campañas activas
		obj_client = Clients.query.filter_by(name=str(client)).first()

	else:
		obj_client = Clients.query.filter_by(id=str(selected_client_id)).first()
		client = obj_client.name
		campaign_id = request.form.get('campaign')

	institutions = Institutions()
	institutions = institutions.get_institutions_not_in_campaign(client = obj_client.id)

	return render_template(
		'campaign_select_institutions.html',
		title = title,
		campaigns_module = campaigns_module,
		institutions = institutions,
		client = client,
		campaign_id = campaign_id
	)

# registro de establecimientos a campaña
@admin_bp.route('/admin/campanias/quitar_establecimiento', methods=['POST'])
@login_required
@admin_required
def campaigns_discard_institution():
	if request.method == 'POST':

		institution_id = request.form.get('institution_id')
		campaign_id = request.form.get('campaign_id')

		try:
			row = InstitutionsCampaigns.query.filter(InstitutionsCampaigns.institution_id == institution_id, campaign_id == campaign_id).first()
			i = Institutions()
			i.finish_insts(institution_id = institution_id, campaign_institution_id = row.id, current_user_id=current_user.id, discard=True)
		except Exception as e:
			print(e)
		finally:
			return redirect(url_for('admin.campaigns_institutions', id = campaign_id, discard = True))

# finalizar campañas
@admin_bp.route('/admin/campanias/terminar', methods=['POST'])
@login_required
@admin_required
def finish_campaign():

	if request.form.get('campaign'):

		id = request.form.get('campaign')
		user_id = request.form.get('user_id')

		institutions = Institutions()
		institutions_in_campaign = institutions.get_institutions_by_campaigns(id)

		institutions_in_campaign = Institutions.query.join(InstitutionsCampaigns).filter(InstitutionsCampaigns.campaign_id == id).all()

		for institution in institutions_in_campaign:
			ic = InstitutionsCampaigns.query.filter(InstitutionsCampaigns.campaign_id == id, InstitutionsCampaigns.institution_id == institution.id).first()
			i = Institutions()
			i.finish_insts(institution_id = institution.id, campaign_institution_id = ic.id, current_user_id=user_id)

		# se finaliza camapaña
		campaign = Campaigns.query.filter_by(id=id).first()
		campaign.id = id
		campaign.finish_campaign()

		lcs = LogCampaignStatus()
		lcs.campaign_id = id
		lcs.user = current_user
		lcs.setLog()

	return redirect(url_for('admin.campaigns'))

# USUARIOS

# lista de usuarios
@admin_bp.route("/admin/usuarios")
@admin_bp.route("/admin/usuarios/")
@login_required
def users():

	title = "Usuarios - Sistema de respaldos"
	users_module = True
	users = Users.query.all()

	return render_template(
		'users.html',
		users_module=users_module,
		title=title,
		users = users)

# registro de usuarios
@admin_bp.route("/admin/usuarios/registrar")
@admin_bp.route("/admin/usuarios/registrar/")
@login_required
def users_register():

	title = "Usuarios - Sistema de respaldos"
	users_module = True
	form = RegisterUsersForm()
	
	return render_template(
		'users_register.html',
		title = title,
		users_module = users_module,
		form = form
	)

@admin_bp.route('/admin/usuarios/signup', methods=["POST"])
@login_required
def user_signup():
	form = RegisterUsersForm()
	error = None
	
	if form.validate_on_submit():
		username = form.username.data
		name = form.name.data
		lastname = form.lastname.data
		email = form.email.data
		password = form.password.data
		category = form.categories.data
		is_admin = form.is_admin.data
		# se compueba si existe usuario con el mismo username
		user = Users.get_by_username(username)
		if user is not None:
			error = f"El nombre de usuario ya existe"
		else:
			user = Users(username=username, name=name, lastname=lastname, email=email, is_admin=is_admin, category = category)
			user.set_password(password)
			user.save()
			
	return redirect(url_for('admin.users'))

# formulario de edicion
@admin_bp.route('/admin/usuarios/editar/<int:id>', methods=["GET", "POST"])
@login_required
def user_edit(id):

	title = "Usuarios - Sistema de respaldos"
	users_module = True
	form = UpdateUserForm()

	if request.method == "GET":
		user = Users.query.filter_by(id=id).first()
		id = user.id
		form.user_id.data = user.id
		form.username.data = user.username
		form.name.data = user.name
		form.lastname.data = user.lastname
		form.email.data = user.email
		form.is_admin.data = user.is_admin
		form.category.data = user.category
		
	return render_template(
		'users_edit.html', 
		form=form, 
		title=title, 
		users_module=users_module,
		id = id)

@admin_bp.route('/admin/usuarios/editar/post', methods=["POST"])
@login_required
def user_edit_post():

	form = UpdateUserForm()
	
	if form.validate_on_submit():
		user = Users.get_by_id(form.user_id.data)
		user.name = form.name.data
		user.username = form.username.data
		user.lastname = form.lastname.data
		user.email = form.email.data
		user.is_admin = form.is_admin.data
		user.category = form.category.data
		user.update()

		return redirect(url_for('admin.users'))
	else:
		print(form.errors)
		
	return redirect(url_for('admin.users'))

@admin_bp.route('/editar_password/<int:id>', methods=["GET"])
@login_required
def user_edit_password(id):

	if request.method == "GET":
		form = UpdatePasswordForm()
		user = Users.get_by_id(str(id))
		form.user_id.data = user.id

	return render_template(
		'users_password.html', 
		title = "Usuarios - Sistema de respaldos", 
		users_module = True, 
		form=form, 
		id=id)

@admin_bp.route('/editar_password/post', methods=["POST"])
@login_required
def user_edit_password_post():

	form = UpdatePasswordForm()

	if form.validate_on_submit():
		if form.password.data == form.repeat_password.data:
			user = Users()
			user.id = str(form.user_id.data)
			user.set_password(form.password.data)
			user.update_password()
			return redirect(url_for('admin.user_edit_password', id = str(form.user_id.data)))
		else:
			return redirect(url_for('admin.user_edit_password', id = str(form.user_id.data)))

	return redirect(url_for('admin.users'))

# lista de tipos de enlaces
@admin_bp.route("/admin/link-types")
@admin_bp.route("/admin/link-types/")
@login_required
def link_types():

	title = "Tipos de Enlaces - Sistema de respaldos"
	link_type_module = True
	link_types = LinkTypes.query.all()

	return render_template(
		'link_types.html',
		link_type_module=link_type_module,
		title=title,
		link_types = link_types)

# registro de tipos de enlaces
@admin_bp.route("/admin/link-types/registrar", methods=['POST','GET'])
@login_required
def link_types_register():

	title = "Tipos de Enlaces - Sistema de respaldos"
	link_type_module = True
	form = RegisterLinkTypesForm()
	error = None

	if form.validate_on_submit():
		name = form.name.data
		links_type = LinkTypes.get_by_name(name)
		if links_type is not None:
			error = f"El tipo de enlace ya existe"
		else:
			link_type = LinkTypes(name=name)
			link_type.save()
			return redirect(url_for('admin.link_types'))
	
	return render_template(
		'link_types_register.html',
		title = title,
		link_type_module = link_type_module,
		form = form,
		error_msg = error

	)


# lista de tipos de conexión
@admin_bp.route("/admin/connection-types")
@admin_bp.route("/admin/connection-types/")
@login_required
def connection_types():

	title = "Tipos de Conexión - Sistema de respaldos"
	connection_type_module = True
	connection_types = ConnectionTypes.query.all()

	return render_template(
		'connection_types.html',
		connection_type_module=connection_type_module,
		title=title,
		connection_types = connection_types)

# registro de tipos de conexión
@admin_bp.route("/admin/connection-types/registrar", methods=['POST','GET'])
@login_required
def connection_types_register():
	title = "Tipos de Conexión - Sistema de respaldos"
	connection_type_module = True
	error= None
	form = RegisterConnectionTypesForm()
	if form.validate_on_submit():
		name = form.name.data
		connection_type = ConnectionTypes.get_by_name(name)
		if connection_type is not None:
			error = f"El tipo de conexión ya existe"
		else:
			connection_type = ConnectionTypes(name=name)
			connection_type.save()
			return redirect(url_for('admin.connection_types'))

	return render_template(
		'connection_types_register.html',
		title = title,
		connection_type_module = connection_type_module,
		form = form,
		error_msg = error
	)

# lista de tipos de enlaces
@admin_bp.route("/admin/enlaces")
@admin_bp.route("/admin/enlaces/<string:link>", methods=["POST"])
@login_required
def links(link = None ):

	title = "Enlaces - Sistema de respaldos"
	link_module = True
	links = Links.query.order_by(Links.id).all()
	link_delete_name = None
	if link is not None:
		link_delete_name = link
	return render_template(
		'links.html',
		link_module=link_module,
		title=title,
		links = links,
		link_delete_name = link_delete_name)

# registro de enlaces
@admin_bp.route("/admin/enlaces/registrar")
@admin_bp.route("/admin/enlaces/registrar/", methods=['POST','GET'])
@admin_bp.route("/admin/enlaces/registrar/<int:client_id>", methods=['POST','GET'])
@login_required
def links_register(client_id=None):
	title = "Enlaces - Sistema de respaldos"
	link_module = True

	form = RegisterLinksForm()
	form.clients.choices=[("","")] +[(client.id, client.name) for client in Clients.query.order_by(Clients.name).all()]
	form.clients.data = str(client_id)

	#tipo de conexión
	form.connection_type.choices = [("","")] + [(ct.id, ct.name) for ct in ConnectionTypes.query.order_by(ConnectionTypes.name).all()]

	#tipo de enlace
	form.link_type.choices = [(lt.id, lt.name) for lt in LinkTypes.query.order_by(LinkTypes.name).all()]

	msg_code_exist = False
	msg = False

	if client_id is None:
		form.institution.choices = []
	else:
		form.institution.choices = [("","")] + [(institution.id, institution.name) for institution in Institutions.query.filter(Institutions.client_id == client_id, Institutions.active == True).order_by(Institutions.name).all()]

	if form.validate_on_submit():
		service_code = str(form.service_code.data)
		ip_direction = str(form.ip_direction.data)
		link_type = str(form.link_type.data)
		connection_type = str(form.connection_type.data)
		institution = form.institution.data

		# código de servicio único por cliente
		sql = Links.query.join(Institutions).filter(Institutions.client_id==form.clients.data, Links.service_code == service_code) \
				.first()
		
		if sql is None:
			if connection_type == '':
				connection_type = None
			links = Links(service_code=service_code,ip_direction = ip_direction, link_type_id= link_type, connection_type_id = connection_type, institution_id = institution)
			links.save()
			msg = True
		else:
			msg_code_exist = True

	return render_template(
		'links_register.html',
		title = title,
		link_module = link_module,
		form = form,
		client_id = client_id,
		msg_code_exist= msg_code_exist,
		msg = msg
	)

# registro de enlaces
@admin_bp.route("/admin/enlaces/editar/<int:id>", methods=['GET','POST'])
@login_required
def links_edit(id):
	title = "Enlaces - Sistema de respaldos"
	link_module = True

	msg = False
	msg_code_exist = False

	form = EditLinksForm()

	link = Links.query.filter_by(id=id).first()

	tmp_service_code = form.service_code.data
	tmp_ip_direction = form.ip_direction.data
	tmp_link_type  = form.link_type.data
	tmp_connection_type = form.connection_type.data

	form.service_code.data = link.service_code
	form.ip_direction.data = link.ip_direction
	form.link_type.data = link.link_type
	form.connection_type.data = link.connection_type
	form.institution.data = link.institution

	if request.method == "POST":
		if form.validate_on_submit():

			# código de servicio único por cliente
			sql = Links.query.join(Institutions).join(Clients) \
				.filter(Clients.id==link.institution.client_id, Links.id != link.id, Links.service_code == tmp_service_code) \
					.first()
			if sql is None:
				link.service_code = tmp_service_code
				link.ip_direction = tmp_ip_direction
				link.link_type = tmp_link_type
				link.connection_type = tmp_connection_type
				link.save()
				msg = True
			else:
				msg_code_exist = True
			
			form.link_id.data = id
			form.service_code.data = tmp_service_code
			form.ip_direction.data = tmp_ip_direction
			form.link_type.data = tmp_link_type
			form.connection_type.data = tmp_connection_type

	return render_template(
		'links_edit.html',
		title = title,
		link_module = link_module,
		form = form,
		id = id,
		msg = msg,
		msg_code_exist = msg_code_exist)

@admin_bp.route("/admin/enlaces/eliminar/<int:id>", methods=["POST"])
@login_required
def links_delete(id):
	if request.method == "POST":
		links = Links()
		link = links.get_by_id(id = id)
		link.delete()
	return redirect(url_for('admin.links', link = link.service_code), code=307)

@admin_bp.route("/admin/enlaces/importar", methods=["GET", "POST"])
@login_required
def links_import():

	form = ImportLinksForm()

	return render_template(
		"links_import.html", 
		form=form,
		title = "Importar enlaces - Sistema de respaldos",
		link_module = True)

@admin_bp.route("/admin/enlaces/importar/preview", methods=["POST"])
@login_required
def links_import_preview():

	form = ImportLinksForm()
	global form_client
	global data_xls_import 
	global tmp_data_import
	count_links = 0

	if form.validate_on_submit():
		filename = secure_filename(form.upload.data.filename)
		
		form_client = form.client.data

		form.upload.data.save(UPLOADS_PATH+filename)

		tmp_data_import = load_workbook(UPLOADS_PATH+filename, data_only=True)
		data_xls_import = pd.read_excel(UPLOADS_PATH+filename, header=0)

		count_links = len(data_xls_import)

		data_xls_import_html = data_xls_import.to_html(classes="table table-bordered")

	return render_template(
		"links_import.html", 
		form=form, 
		data_xls_import=data_xls_import_html, 
		client = form_client, 
		count_links = count_links,
		link_module = True)

# esto se ejecuta al confirmar la importación
# se registran los datos de la planilla a la base de datos
@admin_bp.route("/admin/enlaces/importar/confirmar", methods=["POST"])
@login_required
def links_import_confirm():
	
	global tmp_data_import
	global form_client

	form = ImportLinksForm()

	repeat = True
	count = 1
	num_rows = 0
	count_links_repeat = 0
	institutions_repeated = []
	links_repeated = []
	links_registered = 0
	exist = False
	import_finished = False
	
	no_institution = []
	element_no_institution = []
	no_link_type = []
	element_no_link_type = []
	no_connection_type = []
	limit_ip_address = []
	code_service_exist = []
	none_ip_service_code = []

	while repeat == True:
		num_rows = num_rows + 1
		count = count + 1

		letra_a = "A"+str(count)
		letra_b = "B"+str(count)
		letra_c = "C"+str(count)
		letra_d = "D"+str(count)
		letra_e = "E"+str(count)

		celda_a = tmp_data_import['enlaces'][letra_a].value
		celda_b = tmp_data_import['enlaces'][letra_b].value
		celda_c = tmp_data_import['enlaces'][letra_c].value
		celda_d = tmp_data_import['enlaces'][letra_d].value
		celda_e = tmp_data_import['enlaces'][letra_e].value

		client = Clients()
		client = client.get_id_by_name(name = str(form_client))

		# se valida si quedan más registros
		if (celda_a == None and celda_b == None and celda_c == None):
			repeat = False
		else:
			
			# validación de establecimiento
			institution_exist = Institutions()
			institution_exist = institution_exist.get_institution_by_name(name=str(celda_a), client=client.id)

			link_type_exist = LinkTypes.get_by_name((str(celda_b))) 
			connection_type_exist = ConnectionTypes.get_by_name(str(celda_c))

			l = Links.service_code_exist(str(celda_d))

			# código de servicio único por cliente
			sql = Links.query.join(Institutions).filter(Institutions.client_id==client.id, Links.service_code == str(celda_d)) \
					.first()

			if celda_d is None:
				# código de servicio debe ser obligatorio
				element_none_ip_service_code = {
					'institution': str(celda_a), 
					'link_type': str(celda_b),
					'connection_type': str(celda_c),
					'service_code': str(celda_d),
					'ip_address': str(celda_e)
					}
				none_ip_service_code.append(element_none_ip_service_code)

			elif sql is not None:
				# enlaces con código de servicio ya existentes
				element_code_service_exist = {
					'institution': str(celda_a), 
					'link_type': str(celda_b),
					'connection_type': str(celda_c),
					'service_code': str(celda_d),
					'ip_address': str(celda_e)
					}
				code_service_exist.append(element_code_service_exist)

			elif institution_exist is None:
				# enlaces con establecimientos no encontrados
				element_no_institution = {
					'institution': str(celda_a), 
					'link_type': str(celda_b),
					'connection_type': str(celda_c),
					'service_code': str(celda_d),
					'ip_address': str(celda_e)
					}
				no_institution.append(element_no_institution)

			elif link_type_exist is None:
				# tipo de enlaces no encontrados
				element_no_link_type = {
					'institution': str(celda_a), 
					'link_type': str(celda_b),
					'connection_type': str(celda_c),
					'service_code': str(celda_d),
					'ip_address': str(celda_e)
					}
				no_link_type.append(element_no_link_type)
			# las ip no pueden superar los 15 caracteres
			elif len(str(celda_e)) > 15:
				# tipo de enlaces no encontrados
				element_limit_ip_address = {
					'institution': str(celda_a), 
					'link_type': str(celda_b),
					'connection_type': str(celda_c),
					'service_code': str(celda_d),
					'ip_address': str(celda_e)
					}
				limit_ip_address.append(element_limit_ip_address)

			else:
				links = Links(service_code=str(celda_d),ip_direction = str(celda_e), link_type= link_type_exist, connection_type = connection_type_exist, institution = institution_exist)
				links.save()
				links_registered = links_registered + 1

	import_finished = True

	return render_template(
		"links_import.html",
		client = client,
		form=form, 
		import_finished = import_finished, 
		links_module = True,
		links_registered = links_registered,
		no_institution = no_institution,
		no_link_type = no_link_type,
		limit_ip_address = limit_ip_address,
		code_service_exist = code_service_exist,
		none_ip_service_code = none_ip_service_code)

@admin_bp.route("/admin/correos", methods=["GET"])
@login_required
def mails_list():
	mails = SendMails.query.order_by(SendMails.id).all()
	return render_template(
		"mails_list.html",
		mails = mails,
		mails_module = True
	)

@admin_bp.route("/admin/correos/<string:type>/editar/<int:id>", methods=["GET"])
@admin_bp.route("/admin/correos/<string:type>/editar/<int:id>", methods=["POST"])
@login_required
def mail_cc_edit(id, type):

	contacts = ''
	
	send_mail_descripion = SendMails.query.filter(SendMails.id == id).first()

	users = Users.query.order_by(Users.name).all()
	groupmails = GroupsMails.query.all()

	type_value = 0 if type == 'to' else (1 if type == 'cc' else None)

	# se valida que el grupo no esté registrado para la tarea programada seleccionada
	saved_groups = GroupsMails.query.join(GroupMailsSendMails).filter(GroupMailsSendMails.type == type_value, GroupMailsSendMails.send_mail_id == id).all()

	items_in_groups = ItemGroupsMails.query.all()

	if type == 'cc':
		contacts = SendMailsCC.query.filter(SendMailsCC.send_mail_id == id).join(SendMails).join(Users).all()
		anon_contacts = SendMailsCC.query.filter(SendMailsCC.send_mail_id == id, SendMailsCC.user_id == None).join(SendMails).all()

		if request.method == "POST":
			if request.form.get("anon_user"):
				mail = request.form.get("anon_user")
				new_cc = SendMailsCC(mail = mail, send_mail_id = id)
				new_cc.save()

			if request.form.get("user_id"):
				user_id = request.form.get("user_id")
				new_cc = SendMailsCC(user_id = user_id, send_mail_id = id)
				new_cc.save()

			if request.form.get("remove_email"):
				id_cc = request.form.get("contact_id")
				cc = SendMailsCC.query.filter(SendMailsCC.id == id_cc).first()
				cc.delete()

			if request.form.get("remove_group"):
				group_id = request.form.get("group_id")
				cc = GroupMailsSendMails.query.filter(GroupMailsSendMails.groupmail_id == group_id).first()
				cc.delete()

			return redirect(url_for('admin.mail_cc_edit', id = id, type='cc'))

	if type == 'to':
		contacts = SendMailsFrom.query.filter(SendMailsFrom.send_mail_id == id).join(SendMails).join(Users).all()
		anon_contacts = SendMailsFrom.query.filter(SendMailsFrom.send_mail_id == id, SendMailsFrom.user_id==None).join(SendMails).all()
		if request.method == "POST":
			if request.form.get("anon_user"):
				mail = request.form.get("anon_user")
				new_cc = SendMailsFrom(mail = mail, send_mail_id = id)
				new_cc.save()

			if request.form.get("user_id"):
				user_id = request.form.get("user_id")
				new_cc = SendMailsFrom(user_id = user_id, send_mail_id = id)
				new_cc.save()

			if request.form.get("remove_email"):
				id_cc = request.form.get("contact_id")
				cc = SendMailsFrom.query.filter(SendMailsFrom.id == id_cc).first()
				cc.delete()

			if request.form.get("remove_group"):
				group_id = request.form.get("group_id")
				cc = GroupMailsSendMails.query.filter(GroupMailsSendMails.groupmail_id == group_id).first()
				cc.delete()

			return redirect(url_for('admin.mail_cc_edit', id = id, type='to'))

	return render_template(
		"mails_edit.html",
		contacts = contacts,
		anon_contacts = anon_contacts,
		users = users,
		groupsmails = groupmails,
		saved_groups = saved_groups,
		items_in_groups = items_in_groups,
		send_mail_descripion = send_mail_descripion,
		id = id,
		type = type,
		mails_module = True
	)

@admin_bp.route("/admin/correos/grupos/", methods=['GET','POST'])
@login_required
@admin_required
def groupsmails():

	msg_success = None

	if request.method == "POST":
		idGroup = request.form.get("group_id")

		# eliminar los contactos
		data = ItemGroupsMails.query.filter(ItemGroupsMails.group_mail_id == idGroup).all()
		for item in data:
			item.delete()

		# eliminar relación con los clientes
		data = ClientsGroupsMails.query.filter(ClientsGroupsMails.groupmail_id == idGroup).all()
		for item in data:
			item.delete()

		# eliminar relación con Envíos de correos
		data = GroupMailsSendMails.query.filter(GroupMailsSendMails.groupmail_id == idGroup).all()
		for item in data:
			item.delete()

		# eliminar grupo
		group = GroupsMails.query.filter(GroupsMails.id == idGroup).first()
		groupName = group.name
		group.delete()

		msg_success = "Se ha eliminado el grupo <b>"+groupName+"</b> correctamente."

	groupMails = GroupsMails.query.all()

	return render_template(
		"groupmails.html", 
		title = "Sistema de respaldos - Grupo de Correos", 
		mails_module = True,
		groupMails = groupMails,
		msg_success = msg_success)

@admin_bp.route("/admin/correos/grupos/nuevo", methods=['POST','GET'])
@login_required
def groupsmailsNew():

	form = RegisterGroupsMailsForm()

	if form.validate_on_submit():

		groupmails = GroupsMails(name=form.name.data)
		groupmails.save()

		return redirect(url_for('admin.groupsmails'))

	return render_template(
		"groupmails_register.html", 
		title = "Sistema de respaldos - Grupo de Correos", 
		mails_module = True,
		form = form)

@admin_bp.route("/admin/correos/grupos/<int:id>/editar", methods=['POST','GET'])
@login_required
def groupsmailsEdit(id):

	msg_success = None
	form = EditGroupsMailsForm()
	
	if form.validate_on_submit():
		groupmails = GroupsMails.query.filter(GroupsMails.id == id).first()
		groupmails.name = str(form.name.data)
		groupmails.save()
		form.name.data = groupmails.name
		msg_success = "Se han guardado los cambios exitosamente."

	gp = GroupsMails.query.filter(GroupsMails.id == id).first()
	form.name.data = gp.name

	return render_template(
		"groupmails_edit.html", 
		title = "Sistema de respaldos - Grupo de Correos", 
		mails_module = True,
		form = form,
		id = id,
		msg_success = msg_success)

@admin_bp.route("/admin/correos/grupos/<int:group_id>/agregar", methods=['POST','GET'])
@login_required
def groupsmailsNewItem(group_id):

	error = None

	itemGroupMail = []
	clientsGroupMail = []

	formAddItem = RegisterItemGroupsMailsForm()
	formAddClient = RegisterClientsGroupsMailsForm()

	groupmail = GroupsMails.query.filter(GroupsMails.id == group_id).first()

	if formAddItem.validate_on_submit():
		addItem = ItemGroupsMails(name = formAddItem.name.data, mail = formAddItem.mail.data, group_mail_id = group_id)
		addItem.save()

	if formAddClient.validate_on_submit():

		# se valida que el cliente no pertenezca ya a otro grupo
		tmp = ClientsGroupsMails.query.filter(ClientsGroupsMails.clients == formAddClient.clients.data).first()
		if tmp is None:
			addItem = ClientsGroupsMails(clients = formAddClient.clients.data, groupmail_id = group_id)
			addItem.save()
		else:
			error = 'El cliente seleccionado ya se encuentra registrado en otro grupo.'
			itemGroupMail = ItemGroupsMails.query.filter(ItemGroupsMails.group_mail_id == group_id).all()
			clientsGroupMail = ClientsGroupsMails.query.filter(ClientsGroupsMails.groupmail_id == group_id).join(Clients).all()

			return render_template(
				"groupmails_item_register.html", 
				title = "Sistema de respaldos - Grupo de Correos", 
				mails_module = True,
				formAddItem = formAddItem,
				groupmail = groupmail,
				itemGroupMail = itemGroupMail,
				formAddClient = formAddClient,
				clientsGroupMail = clientsGroupMail,
				error = error,
			)

	itemGroupMail = ItemGroupsMails.query.filter(ItemGroupsMails.group_mail_id == group_id).all()
	clientsGroupMail = ClientsGroupsMails.query.filter(ClientsGroupsMails.groupmail_id == group_id).join(Clients).all()

	return render_template(
		"groupmails_item_register.html", 
		title = "Sistema de respaldos - Grupo de Correos", 
		mails_module = True,
		formAddItem = formAddItem,
		groupmail = groupmail,
		itemGroupMail = itemGroupMail,
		formAddClient = formAddClient,
		clientsGroupMail = clientsGroupMail
	)

@admin_bp.route("/admin/correos/grupos/delete/<int:type>/<int:group_id>/<int:id>", methods=['GET'])
@login_required
def groupsmailsDeleteItem(type, group_id, id):

	# type = 1 = item
	# type = 2 = client

	if type == 1:
		item = ItemGroupsMails.query.filter(ItemGroupsMails.id == id).first()
		item.delete()
	elif type == 2:
		item = ClientsGroupsMails.query.filter(ClientsGroupsMails.id == id).first()
		item.delete()

	return redirect(url_for('admin.groupsmailsNewItem', group_id = group_id))

@admin_bp.route("/admin/correos/groups/<string:type>/<int:sendmail_id>/post", methods=['POST'])
@login_required
def groupsmails_add(sendmail_id, type):

	if request.method == "POST":

		groupmail_id = request.form.get("groupmail_id")

		type_value = 0 if type == 'to' else (1 if type == 'cc' else None)

		# se valida que el grupo no esté registrado para la tarea programada seleccionada
		row = GroupMailsSendMails.query.filter(GroupMailsSendMails.groupmail_id == groupmail_id, GroupMailsSendMails.send_mail_id == sendmail_id).first()

		if row is None:
			try:
				obj = GroupMailsSendMails(groupmail_id = groupmail_id, send_mail_id = sendmail_id, type = type_value)
				obj.save()

			except Exception as e:
				print(e)
		
		return redirect(url_for('admin.mail_cc_edit', id = sendmail_id, type=type))

@admin_bp.route("/historial/establecimientos", methods=["GET","POST"])
@login_required
@admin_required
def logs_institutions():

	period = None

	if request.form.get('period'):
		period = request.form.get('period')

	logs = LogInstitutionsStatus.query.join(InstitutionsCampaigns).join(Institutions).join(Campaigns).filter(Campaigns.period == period).order_by(LogInstitutionsStatus.created.desc()).all()

	periods = db.session.query(Campaigns.period).group_by(Campaigns.period)

	return render_template(
		"logs_institutions.html",
		institutions = logs,
		menu_logs = True,
		menu_logs_institutions = True,
		periods = periods
	)


@admin_bp.route("/historial/campañas", methods=["GET"])
@login_required
@admin_required
def logs_campaigns():
	logs = LogCampaignStatus.query.join(Campaigns).join(Users).order_by(LogCampaignStatus.created.desc()).all()
	return render_template(
		"logs_campaigns.html",
		campaigns = logs,
		menu_logs = True,
		menu_logs_campaigns = True
	)